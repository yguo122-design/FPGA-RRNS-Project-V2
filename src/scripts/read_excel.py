import openpyxl
wb = openpyxl.load_workbook("docs/dissertation/figure/power latency resource summary.xlsx")
for sheet in wb.sheetnames:
    print(f"Sheet: {sheet}")
    ws = wb[sheet]
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            print(row)
